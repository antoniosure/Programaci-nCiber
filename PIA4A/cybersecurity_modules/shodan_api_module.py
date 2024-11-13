"""
Este modulo realiza busquedas com Api shodan, por lo cual necesita contar
con su api key, el cual deberá tener en un archivo .txt, para el archivo
donde se guardará sus resultados deberá escribir la ruta completa del archivo.

El término de búsqueda en este contexto se refiere a la palabra 
clave o frase que quieres buscar en la base de datos de Shodan. 
Shodan es un motor de búsqueda para dispositivos y servicios 
conectados a internet, como cámaras IP, servidores web, routers, y otros sistemas. 

Algunos ejemplos de términos de búsqueda comunes en Shodan podrían ser:

1. "apache"- para encontrar servidores que usan Apache HTTP.
2. "default password"- para localizar dispositivos con configuraciones de contraseña por defecto.
3. "webcam" o "camera"- para buscar cámaras de seguridad conectadas a internet.
4. "port:22"- para buscar dispositivos con el puerto 22 abierto (comúnmente usado por SSH).
5. "country:US" o "country:MX"- para limitar la búsqueda a un país específico 
   (en este caso, EE.UU. o México).

Cuando en el menú elijas "Realizar búsqueda en Shodan" y se te pida el "término de búsqueda", 
puedes ingresar cualquiera de estos ejemplos o el término que sea de tu 
interés para obtener datos relevantes en la búsqueda de Shodan.
"""

import os
import time
import shodan
import sys
import hashlib
import platform
import subprocess
from openpyxl import Workbook, load_workbook

def check_os():
    os_type = platform.system()
    if os_type not in ["Windows", "Linux"]:
        print("Este script solo es compatible con Windows o Linux.")
        exit()
    return os_type

def check_powershell():
    try:
        subprocess.run(["powershell", "-Command", "Write-Output 'Powershell está instalado'"], check=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
        print("Powershell está instalado.")
        return True
    except (subprocess.CalledProcessError, FileNotFoundError):
        print("Powershell no está instalado.")
        return False

def install_powershell_linux():
    try:
        subprocess.run(["sudo", "apt-get", "update"], check=True)
        subprocess.run(["sudo", "apt-get", "install", "-y", "powershell"], check=True)
        print("Powershell instalado con éxito en Linux.")
    except subprocess.CalledProcessError:
        print("Error al instalar Powershell. Por favor, instálalo manualmente.")
        exit()

def check_and_install_shodan():
    try:
        import shodan
        print("La API de Shodan está instalada.")
    except ImportError:
        print("La API de Shodan no está instalada. Procediendo a instalarla.")
        subprocess.check_call([sys.executable, "-m", "pip", "install", "shodan"])
        print("API de Shodan instalada con éxito.")

def read_api_key(filepath):
    try:
        with open(filepath, 'r') as file:
            api_key = file.read().strip()
            print("Clave de API cargada exitosamente.")
            return api_key
    except FileNotFoundError:
        print("Error: archivo no encontrado. Verifica la ruta e inténtalo de nuevo.")
        return None

def save_to_file(data, output_file):
    with open(output_file, 'a') as file:
        file.write(data)
    print(f"Datos guardados en {output_file}")

def calculate_hash(file_path):
    # Calcula el hash SHA-256 de un archivo en lugar de una cadena
    hash_sha256 = hashlib.sha256()
    try:
        with open(file_path, "rb") as f:
            for chunk in iter(lambda: f.read(4096), b""):
                hash_sha256.update(chunk)
    except FileNotFoundError:
        print("El archivo no existe para calcular el hash.")
        return None
    return hash_sha256.hexdigest()

def shodan_search(api_key, query, output_file):
    api = shodan.Shodan(api_key)
    counter = 0
    limit = 10

    try:
        print("Autenticando clave de API en Shodan...")
        api.search("test")
        print("Autenticación exitosa.")
        print(f"Iniciando búsqueda para el término: {query}")
        for banner in api.search_cursor(query):
            data = (
                f"IP: {banner['ip_str']}\n"
                f"Port: {banner['port']}\n"
                f"Organization: {banner.get('org', 'N/A')}\n"
                f"Location: {banner.get('location', 'N/A')}\n"
                f"Layer: {banner['transport']}\n"
                f"Domains: {banner.get('domains', 'N/A')}\n"
                f"Hostnames: {banner.get('hostnames', 'N/A')}\n"
                f"Data: {banner['data']}\n"
                "-----------------------------------\n"
            )
            save_to_file(data, output_file)
            counter += 1
            if counter >= limit:
                print("Límite de resultados alcanzado.")
                break
    except shodan.APIError as e:
        print(f"Error de API de Shodan: {e}")
    except KeyboardInterrupt:
        print("Proceso interrumpido por el usuario.")
    
    # Al finalizar la búsqueda, calcular y guardar el hash del archivo de salida
    output_file_hash = calculate_hash(output_file)
    if output_file_hash:
        save_output_file_hash_to_excel(output_file, output_file_hash)

def save_output_file_hash_to_excel(output_file, output_file_hash):
    ruta_excel = os.path.join(os.path.expanduser("~"), "Documents", "registro_shodan_archivos.xlsx")

    # Verificar si el archivo de Excel ya existe
    if os.path.exists(ruta_excel):
        wb = load_workbook(ruta_excel)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(["Archivo de salida", "Hash SHA-256", "Fecha de creación"])  # Encabezados

    # Agregar registro al Excel
    timestamp = time.strftime("%Y-%m-%d %H:%M:%S")
    ws.append([output_file, output_file_hash, timestamp])

    wb.save(ruta_excel)
    print(f"Registro del archivo de salida guardado en {ruta_excel}")

def main_menu():
    os_type = check_os()
    if os_type == "Linux" and not check_powershell():
        install_powershell_linux()
    elif os_type == "Windows" and not check_powershell():
        print("Instale PowerShell manualmente en Windows para ejecutar este script.")
        exit()

    check_and_install_shodan()

    api_key = None
    output_file = None

    while True:
        print("\n--- Menú Interactivo Shodan ---")
        print("1. Cargar clave de API desde archivo")
        print("2. Configurar archivo de salida")
        print("3. Realizar búsqueda en Shodan")
        print("4. Salir")
        choice = input("Selecciona una opción: ")

        if choice == "1":
            filepath = input("Ingresa la ruta del archivo de la clave de API: ")
            api_key = read_api_key(filepath)
        elif choice == "2":
            output_file = input("Ingresa el nombre del archivo de salida (.txt): ")
        elif choice == "3":
            if api_key is None:
                print("Primero debes cargar una clave de API.")
            elif output_file is None:
                print("Primero debes configurar un archivo de salida.")
            else:
                query = input("Ingresa el término de búsqueda: ")
                shodan_search(api_key, query, output_file)
        elif choice == "4":
            print("Saliendo del programa.")
            break
        else:
            print("Opción no válida. Inténtalo de nuevo.")

# Ejecutar desde este archivo
#if __name__ == "__main__":
    #main_menu()
